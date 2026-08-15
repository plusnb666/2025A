"""CUMCM 2025A 结果填写模块 —— 把求解结果写入 result1/2/3.xlsx。

约定（建模思路.md §7）：多球合力(B)下，每弹"有效干扰时长"填其单独完全遮蔽的时长（名义值）。
"""
import numpy as np
from openpyxl import load_workbook
from model import drone_pos, detonation, cylinder_surface, hit_time
from problems import shield_all

_FINE = cylinder_surface(48, 4)


def drop_det(info):
    """info=(drone_id, theta, v, t_r, dt) → (drop(3,), det(3,), t_d)"""
    drone_id, theta, v, t_r, dt_ = info
    drop = drone_pos(drone_id, theta, v, t_r)
    det, t_d = detonation(drone_id, theta, v, t_r, dt_)
    return drop, det, t_d


def ind_time(k, info, dt=0.001):
    drop, det, t_d = drop_det(info)
    T, _, _, _ = shield_all(k, [(det, t_d)], _FINE, dt, hit_time(k))
    return T


def _w(ws, r, c, v, n=3):
    if v is None:
        ws.cell(r, c).value = None
    else:
        ws.cell(r, c, float(np.round(float(v), n)))


def fill_result1(infos):
    """infos: 3 个 (drone_id, theta, v, t_r, dt)，FY1 对 M1。"""
    wb = load_workbook('result1.xlsx')
    ws = wb.active
    for r, info in enumerate(infos, start=2):
        drop, det, _ = drop_det(info)
        T = ind_time(1, info)
        _w(ws, r, 1, info[1], 2)          # 方向
        _w(ws, r, 2, info[2], 2)          # 速度
        _w(ws, r, 4, drop[0], 1); _w(ws, r, 5, drop[1], 1); _w(ws, r, 6, drop[2], 1)
        _w(ws, r, 7, det[0], 1); _w(ws, r, 8, det[1], 1); _w(ws, r, 9, det[2], 1)
        _w(ws, r, 10, T, 3)               # 有效时长
    wb.save('result1.xlsx')
    print('result1.xlsx 已写入')


def fill_result2(infos):
    """infos: 3 个 (drone_id, theta, v, t_r, dt)，对应 FY1/FY2/FY3 对 M1。"""
    wb = load_workbook('result2.xlsx')
    ws = wb.active
    for r, info in enumerate(infos, start=2):
        drop, det, _ = drop_det(info)
        T = ind_time(1, info)
        _w(ws, r, 2, info[1], 2)
        _w(ws, r, 3, info[2], 2)
        _w(ws, r, 4, drop[0], 1); _w(ws, r, 5, drop[1], 1); _w(ws, r, 6, drop[2], 1)
        _w(ws, r, 7, det[0], 1); _w(ws, r, 8, det[1], 1); _w(ws, r, 9, det[2], 1)
        _w(ws, r, 10, T, 3)
    wb.save('result2.xlsx')
    print('result2.xlsx 已写入')


def fill_result3(entries):
    """entries: list of (drone_id, bomb_idx, info, missile_no)。
    bomb_idx ∈ {1,2,3} 为该无人机内投放顺序；模板行 = 2 + (drone_id-1)*3 + (bomb_idx-1)。"""
    wb = load_workbook('result3.xlsx')
    ws = wb.active
    # 先清空数据区（列 2,3,5..12）
    for r in range(2, 17):
        for c in (2, 3, 5, 6, 7, 8, 9, 10, 11, 12):
            ws.cell(r, c).value = None
    for drone_id, bomb_idx, info, mk in entries:
        r = 2 + (drone_id - 1) * 3 + (bomb_idx - 1)
        drop, det, _ = drop_det(info)
        T = ind_time(mk, info)
        _w(ws, r, 2, info[1], 2)          # 方向
        _w(ws, r, 3, info[2], 2)          # 速度
        _w(ws, r, 5, drop[0], 1); _w(ws, r, 6, drop[1], 1); _w(ws, r, 7, drop[2], 1)
        _w(ws, r, 8, det[0], 1); _w(ws, r, 9, det[1], 1); _w(ws, r, 10, det[2], 1)
        _w(ws, r, 11, T, 3)               # 有效时长
        ws.cell(r, 12, mk)                # 干扰的导弹编号
    wb.save('result3.xlsx')
    print('result3.xlsx 已写入')
