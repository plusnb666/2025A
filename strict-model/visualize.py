"""CUMCM 2025A strict-model 可视化 —— 生成 figures/ 下的 PNG 图。

图 1  作战态势图（俯视 x-y）
图 2  Q1 遮蔽过程（烟幕→视线距离随时间 + 遮蔽窗口）
图 3  各问题结果汇总（含 Q5 三导弹分解）
图 4  代码模块结构与数据流

配色：Okabe-Ito 色盲安全色板（分类色固定顺序，不用彩虹色）。
"""
import os
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.patches import Circle, FancyBboxPatch, FancyArrowPatch

# 注册中文字体（Noto Sans CJK SC）
_CJK = '/usr/share/fonts/google-noto-sans-cjk-fonts/NotoSansCJK-Regular.ttc'
for _f in font_manager.findSystemFonts():
    if 'NotoSansCJK' in _f:
        try:
            font_manager.fontManager.addfont(_f)
        except Exception:
            pass
plt.rcParams['font.family'] = 'Noto Sans CJK SC'
plt.rcParams['axes.unicode_minus'] = False

from model import (MISSILES, DRONES, T_CX, T_CY, T_R, T_H, R_SMOKE,
                   missile_pos, missile_dir, detonation, smoke_centers,
                   seg_dist, hit_time)

HERE = os.path.dirname(os.path.abspath(__file__))
FIG = os.path.join(HERE, 'figures')
os.makedirs(FIG, exist_ok=True)

# Okabe-Ito 色盲安全色板
C = dict(M1='#0072B2', M2='#E69F00', M3='#009E73',
         target='#D55E00', decoy='#7f7f7f', smoke='#CC79A7',
         drone='#56B4E9', threshold='#111111')


def _style(ax):
    ax.grid(True, color='#e3e3e3', linewidth=0.6, alpha=0.7)
    for s in ('top', 'right'):
        ax.spines[s].set_visible(False)
    ax.tick_params(labelsize=9)


# ----------------------------------------------------------------------
# 图 1：作战态势图（俯视 x-y 平面）
# ----------------------------------------------------------------------
def fig_geometry():
    fig, ax = plt.subplots(figsize=(8, 4.6), dpi=150)

    # 导弹轨迹（指向原点）
    for k, col in ((1, C['M1']), (2, C['M2']), (3, C['M3'])):
        P = MISSILES[k]
        ax.plot([P[0], 0], [P[1], 0], color=col, lw=1.6, alpha=0.85)
        ax.annotate(f'M{k}', (P[0], P[1]), color=col, fontsize=9, fontweight='bold',
                    xytext=(8, 4), textcoords='offset points')
    # 无人机
    for i in range(1, 6):
        x, y, _ = DRONES[i]
        ax.plot(x, y, 's', color=C['drone'], ms=5, mec='white', mew=0.5)
        ax.annotate(f'FY{i}', (x, y), color='#444444', fontsize=7,
                    xytext=(0, 8), textcoords='offset points', ha='center')
    # 假目标（原点）
    ax.plot(0, 0, '^', color=C['decoy'], ms=7, mec='white', mew=0.5)
    ax.annotate('假目标(原点)', (0, 0), color='#555555', fontsize=8,
                xytext=(8, -14), textcoords='offset points')
    # 真目标圆柱（俯视圆）
    ax.add_patch(Circle((T_CX, T_CY), T_R, fill=False, color=C['target'],
                        lw=1.8, ls='--'))
    ax.plot(T_CX, T_CY, 'o', color=C['target'], ms=5)
    ax.annotate('真目标\nr=7m', (T_CX, T_CY), color=C['target'], fontsize=8,
                xytext=(10, 6), textcoords='offset points')

    ax.set_xlabel('x (m)'); ax.set_ylabel('y (m)')
    ax.set_title('作战态势（俯视图，导弹直指假目标原点）', fontsize=11)
    ax.set_aspect('equal')
    _style(ax)
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, 'fig1_geometry.png'), bbox_inches='tight')
    plt.close(fig)


# ----------------------------------------------------------------------
# 图 2：Q1 遮蔽过程（烟幕→视线距离随时间）
# ----------------------------------------------------------------------
def fig_q1_process():
    # Q1 固定参数
    theta, v, t_r, dt = 180.0, 120.0, 1.5, 3.6
    det, t_d = detonation(1, theta, v, t_r, dt)          # 起爆点、时刻
    target = np.array([T_CX, T_CY, 5.0])                 # 轴中点

    t = np.arange(t_d, t_d + 20.0, 0.001)
    M = missile_pos(1, t)                                # (N,3)
    Cc = smoke_centers(det, t_d, t)                      # 烟幕球心 (N,3)
    d = seg_dist(Cc, M, target)                          # 球心到视线(导弹→目标)距离

    fig, ax = plt.subplots(figsize=(8, 4.2), dpi=150)
    ax.plot(t, d, color=C['M1'], lw=1.6)
    ax.axhline(R_SMOKE, color=C['threshold'], lw=1.2, ls='--')
    ax.text(t[-1], R_SMOKE + 0.5, ' 遮蔽阈值 10 m', fontsize=8, va='bottom',
            color=C['threshold'])

    # 遮蔽窗口（距离 ≤ 10 m）
    mask = d <= R_SMOKE
    if mask.any():
        ax.axvspan(t[mask].min(), t[mask].max(), color=C['smoke'], alpha=0.25)
        ax.text(t[mask].min(), 0.5, f' 遮蔽窗口 [{t[mask].min():.2f}, {t[mask].max():.2f}] s',
                fontsize=8, color=C['smoke'], va='bottom')

    ax.set_xlabel('时间 t (s)')
    ax.set_ylabel('烟幕球心 → 视线距离 (m)')
    ax.set_title('Q1：烟幕对 M1 视线的遮蔽过程', fontsize=11)
    ax.set_ylim(0, d.max() * 1.05)
    _style(ax)
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, 'fig2_q1_shielding.png'), bbox_inches='tight')
    plt.close(fig)


# ----------------------------------------------------------------------
# 图 3：各问题结果汇总
# ----------------------------------------------------------------------
def fig_results():
    q = ['Q1', 'Q2', 'Q3', 'Q4', 'Q5(Σ)']
    val = [1.392, 4.52, 4.35, 8.43, 27.63]
    fig, ax = plt.subplots(figsize=(7, 3.8), dpi=150)
    bars = ax.bar(q, val, color=[C['M1'], C['M2'], C['M3'], C['smoke'], C['target']],
                  width=0.6, edgecolor='white')
    for b, v in zip(bars, val):
        ax.text(b.get_x() + b.get_width() / 2, v + 0.4, f'{v:.2f}',
                ha='center', fontsize=9)
    ax.set_ylabel('有效遮蔽时长 (s)')
    ax.set_title('各问题有效遮蔽时长（严格完整圆柱判据）', fontsize=11)
    ax.set_ylim(0, max(val) * 1.15)
    _style(ax)
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, 'fig3_results.png'), bbox_inches='tight')
    plt.close(fig)


# ----------------------------------------------------------------------
# 图 4：代码模块结构与数据流
# ----------------------------------------------------------------------
def fig_code_structure():
    fig, ax = plt.subplots(figsize=(8, 5.2), dpi=150)
    ax.axis('off')

    def box(x, y, w, h, text, fc):
        ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle='round,pad=0.02',
                                    fc=fc, ec='#555555', lw=1.0))
        ax.text(x + w / 2, y + h / 2, text, ha='center', va='center', fontsize=9)

    def arrow(x1, y1, x2, y2):
        ax.add_patch(FancyArrowPatch((x1, y1), (x2, y2), arrowstyle='-|>',
                                     mutation_scale=13, color='#555555', lw=1.1))

    # 布局
    box(0.30, 0.82, 0.40, 0.10, 'main.py\n主流程', '#dbe9f4')
    box(0.05, 0.42, 0.28, 0.18, 'model.py\n物理运动学\n+ 几何判定', '#eaf3ea')
    box(0.36, 0.42, 0.28, 0.18, 'problems.py\nQ2~Q5 优化\n(DE + 代理目标)', '#fdf3e3')
    box(0.67, 0.42, 0.28, 0.18, 'fill_result.py\n写 xlsx', '#f6e8f2')
    box(0.67, 0.12, 0.28, 0.10, 'result1/2/3.xlsx', '#f0f0f0')

    arrow(0.42, 0.82, 0.19, 0.60)
    arrow(0.50, 0.82, 0.50, 0.60)
    arrow(0.58, 0.82, 0.81, 0.60)
    arrow(0.19, 0.42, 0.50, 0.42)
    arrow(0.64, 0.42, 0.81, 0.42)
    arrow(0.81, 0.42, 0.81, 0.22)

    ax.text(0.50, 0.97, '代码模块结构与数据流', ha='center', fontsize=12, fontweight='bold')
    ax.text(0.19, 0.33, '导入', ha='center', fontsize=7, color='#666')
    ax.text(0.81, 0.24, '写入', ha='center', fontsize=7, color='#666')
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, 'fig4_code_structure.png'), bbox_inches='tight')
    plt.close(fig)


# ----------------------------------------------------------------------
# 图 5：三维态势图（导弹轨迹 + 无人机 + 真目标圆柱 + Q4 最优云团）
# ----------------------------------------------------------------------
def fig_geometry_3d():
    import openpyxl
    from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

    fig = plt.figure(figsize=(10.5, 7.8), dpi=150)
    ax = fig.add_subplot(111, projection='3d')

    # 导弹轨迹（起点 → 假目标原点）
    for k, col in ((1, C['M1']), (2, C['M2']), (3, C['M3'])):
        P = MISSILES[k]
        ax.plot([P[0], 0], [P[1], 0], [P[2], 0], color=col, lw=1.6)
        ax.text(P[0], P[1], P[2] + 120, f'M{k}', color=col, fontsize=9, fontweight='bold')

    # 无人机初始位置
    for i in range(1, 6):
        x, y, z = DRONES[i]
        ax.scatter([x], [y], [z], s=28, facecolor='white', edgecolor='#555555', lw=1.1)
        ax.text(x, y, z + 120, f'FY{i}', color='#444444', fontsize=8)

    # 真目标圆柱（r=7, h=10, 底面圆心 (0,200,0)）
    ang = np.linspace(0, 2 * np.pi, 48)
    zz = np.array([0.0, T_H])
    A, Z = np.meshgrid(ang, zz)
    X = T_R * np.cos(A)
    Y = T_CY + T_R * np.sin(A)
    ax.plot_surface(X, Y, Z, color=C['target'], alpha=0.65, edgecolor='none')
    for zc in (0.0, T_H):
        ax.plot(T_R * np.cos(ang), T_CY + T_R * np.sin(ang), np.full_like(ang, zc),
                color=C['target'], lw=0.8)
    ax.text(0, 200, 110, '真目标', color=C['target'], fontsize=9, fontweight='bold')

    # 假目标（原点）
    ax.scatter([0], [0], [0], marker='x', s=60, color='black', lw=2)
    ax.text(0, -520, 130, '假目标(原点)', color='black', fontsize=8)

    # Q4 最优云团 + 投放轨迹（从 result2.xlsx 读投放点/起爆点）
    wb = openpyxl.load_workbook(os.path.join(HERE, 'result2.xlsx'))
    ws = wb.active
    for r in range(2, 5):
        drop = np.array([ws.cell(r, c).value for c in (4, 5, 6)], dtype=float)
        det = np.array([ws.cell(r, c).value for c in (7, 8, 9)], dtype=float)
        ax.plot([drop[0], det[0]], [drop[1], det[1]], [drop[2], det[2]],
                '--', color=C['smoke'], lw=1.3)
        u, vv = np.mgrid[0:2 * np.pi:24j, 0:np.pi:12j]
        xs = det[0] + R_SMOKE * np.cos(u) * np.sin(vv)
        ys = det[1] + R_SMOKE * np.sin(u) * np.sin(vv)
        zs = det[2] + R_SMOKE * np.cos(vv)
        ax.plot_surface(xs, ys, zs, color=C['smoke'], alpha=0.35, edgecolor='none')
        ax.scatter([det[0]], [det[1]], [det[2]], color=C['smoke'], s=16)

    ax.set_xlabel('x (m)'); ax.set_ylabel('y (m)'); ax.set_zlabel('z (m)')
    ax.set_title('烟幕干扰弹投放场景三维图（含 Q4 最优云团）', fontsize=12)
    ax.set_xlim(0, 21000); ax.set_ylim(-3500, 2500); ax.set_zlim(0, 2300)
    ax.view_init(elev=22, azim=-58)
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, 'fig5_geometry_3d.png'), bbox_inches='tight')
    plt.close(fig)


# ----------------------------------------------------------------------
# 图 6：Q1 动态演示（GIF：导弹飞行 + 烟幕成形下沉 + 视线遮蔽变色）
# ----------------------------------------------------------------------
def _target_cylinder(ax):
    """在 ax 上画真目标圆柱 + 假目标原点（静态元素）。"""
    ang = np.linspace(0, 2 * np.pi, 48)
    zz = np.array([0.0, T_H])
    A, Z = np.meshgrid(ang, zz)
    X = T_R * np.cos(A)
    Y = T_CY + T_R * np.sin(A)
    ax.plot_surface(X, Y, Z, color=C['target'], alpha=0.65, edgecolor='none')
    for zc in (0.0, T_H):
        ax.plot(T_R * np.cos(ang), T_CY + T_R * np.sin(ang), np.full_like(ang, zc),
                color=C['target'], lw=0.8)
    ax.text(0, 200, 110, '真目标', color=C['target'], fontsize=9, fontweight='bold')
    ax.scatter([0], [0], [0], marker='x', s=60, color='black', lw=2)
    ax.text(0, -520, 130, '假目标(原点)', color='black', fontsize=8)


def animate_q1():
    from matplotlib.animation import FuncAnimation, PillowWriter

    theta, v, t_r, dt = 180.0, 120.0, 1.5, 3.6
    det, t_d = detonation(1, theta, v, t_r, dt)
    target = np.array([T_CX, T_CY, 5.0])

    t_frames = np.linspace(0, 25.0, 150)
    M_all = missile_pos(1, t_frames)                 # (150, 3)

    # 预计算每帧是否遮蔽（烟幕球心到视线距离 ≤ 10 m）
    blocked = np.zeros(len(t_frames), dtype=bool)
    for i, t in enumerate(t_frames):
        if t >= t_d:
            Cc = det + np.array([0.0, 0.0, -3.0]) * (t - t_d)   # (3,)
            blocked[i] = float(seg_dist(Cc, M_all[i], target)) <= R_SMOKE

    fig = plt.figure(figsize=(10, 7.5), dpi=110)
    ax = fig.add_subplot(111, projection='3d')

    def draw(i):
        ax.clear()
        t = t_frames[i]
        M = M_all[i]
        # 导弹全轨迹（淡）+ 当前点
        P0 = MISSILES[1]
        ax.plot([P0[0], 0], [P0[1], 0], [P0[2], 0], color=C['M1'], lw=1.0, alpha=0.25)
        ax.scatter([M[0]], [M[1]], [M[2]], color=C['M1'], s=40)
        ax.text(M[0], M[1], M[2] + 150, 'M1', color=C['M1'], fontsize=9, fontweight='bold')

        _target_cylinder(ax)

        # 视线：遮蔽时高亮，否则淡虚线
        if blocked[i]:
            ax.plot([M[0], target[0]], [M[1], target[1]], [M[2], target[2]],
                    color=C['smoke'], lw=2.2)
        else:
            ax.plot([M[0], target[0]], [M[1], target[1]], [M[2], target[2]],
                    color='#bbbbbb', lw=1.0, ls=':')

        # 烟幕球（起爆后出现，下沉）
        if t >= t_d:
            Ccx, Ccy, Ccz = det + np.array([0.0, 0.0, -3.0]) * (t - t_d)
            u, vv = np.mgrid[0:2 * np.pi:16j, 0:np.pi:8j]
            xs = Ccx + R_SMOKE * np.cos(u) * np.sin(vv)
            ys = Ccy + R_SMOKE * np.sin(u) * np.sin(vv)
            zs = Ccz + R_SMOKE * np.cos(vv)
            ax.plot_surface(xs, ys, zs, color=C['smoke'], alpha=0.5, edgecolor='none')

        status = '遮蔽中' if blocked[i] else ('烟幕已起爆' if t >= t_d else '烟幕未起爆')
        ax.set_title(f'Q1 动态演示  t={t:.1f}s  {status}', fontsize=12)
        ax.set_xlim(0, 21000); ax.set_ylim(-3500, 2500); ax.set_zlim(0, 2300)
        ax.view_init(elev=22, azim=-58)
        ax.set_xlabel('x (m)'); ax.set_ylabel('y (m)'); ax.set_zlabel('z (m)')

    ani = FuncAnimation(fig, draw, frames=len(t_frames), interval=80)
    ani.save(os.path.join(FIG, 'fig6_q1_animation.gif'), writer=PillowWriter(fps=10))
    plt.close(fig)


if __name__ == '__main__':
    fig_geometry()
    fig_q1_process()
    fig_results()
    fig_code_structure()
    fig_geometry_3d()
    animate_q1()
    print('图已生成到', FIG)
    for f in sorted(os.listdir(FIG)):
        print('  ', f)
