# -*- coding: utf-8 -*-
"""
write_xlsx.py —— 把求解结果写回官方模板 result1/2/3.xlsx。

模板已含表头与预填列（弹编号/无人机编号），本模块只填数据列。
"""
import os
import core
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 项目根目录


def _drop(b):
    return core.uav_pos(b['uav'], b['theta'], b['v'], b['t_drop'])


def _det(b):
    return core.bomb_pos(b['uav'], b['theta'], b['v'], b['t_drop'], b['delta'])


def _duration(b, missile):
    det = _det(b)
    t_det = b['t_drop'] + b['delta']
    return core.cloud_blocking_time(det, t_det, missile, dt=1e-4)


def _r(x, nd=2):
    return round(float(x), nd)


def write_result1(bombs, out=None):
    """Q3：FY1 三弹 -> result1.xlsx。bombs 已按投放顺序排列。"""
    src = os.path.join(BASE, 'result1.xlsx')
    out = out or os.path.join(BASE, 'result1_out.xlsx')
    wb = load_workbook(src)
    ws = wb.active
    for i, b in enumerate(bombs):
        row = 2 + i
        drop, det = _drop(b), _det(b)
        ws.cell(row, 1, _r(b['theta']))            # 方向
        ws.cell(row, 2, _r(b['v']))                # 速度
        # C 列弹编号已预填
        ws.cell(row, 4, _r(drop[0])); ws.cell(row, 5, _r(drop[1])); ws.cell(row, 6, _r(drop[2]))
        ws.cell(row, 7, _r(det[0])); ws.cell(row, 8, _r(det[1])); ws.cell(row, 9, _r(det[2]))
        ws.cell(row, 10, _r(_duration(b, 'M1'), 4))  # 有效干扰时长
    wb.save(out)
    return out


def write_result2(bombs, out=None):
    """Q4：FY1/FY2/FY3 各一弹 -> result2.xlsx。bombs 按无人机名顺序。"""
    src = os.path.join(BASE, 'result2.xlsx')
    out = out or os.path.join(BASE, 'result2_out.xlsx')
    wb = load_workbook(src)
    ws = wb.active
    for i, b in enumerate(bombs):
        row = 2 + i
        drop, det = _drop(b), _det(b)
        # A 列无人机编号已预填
        ws.cell(row, 2, _r(b['theta']))
        ws.cell(row, 3, _r(b['v']))
        ws.cell(row, 4, _r(drop[0])); ws.cell(row, 5, _r(drop[1])); ws.cell(row, 6, _r(drop[2]))
        ws.cell(row, 7, _r(det[0])); ws.cell(row, 8, _r(det[1])); ws.cell(row, 9, _r(det[2]))
        ws.cell(row, 10, _r(_duration(b, 'M1'), 4))
    wb.save(out)
    return out


def write_result3(bombs, out=None):
    """Q5：15 弹 -> result3.xlsx。bombs 需含 'missile' 字段；按模板 UAV 顺序排布。"""
    src = os.path.join(BASE, 'result3.xlsx')
    out = out or os.path.join(BASE, 'result3_out.xlsx')
    wb = load_workbook(src)
    ws = wb.active

    uav_order = ['FY1', 'FY2', 'FY3', 'FY4', 'FY5']
    ordered = []
    for u in uav_order:
        group = sorted([b for b in bombs if b['uav'] == u], key=lambda b: b['t_drop'])
        ordered.extend(group)

    for i, b in enumerate(ordered):
        row = 2 + i
        drop, det = _drop(b), _det(b)
        # A 无人机编号、D 弹编号已预填
        ws.cell(row, 2, _r(b['theta']))
        ws.cell(row, 3, _r(b['v']))
        ws.cell(row, 5, _r(drop[0])); ws.cell(row, 6, _r(drop[1])); ws.cell(row, 7, _r(drop[2]))
        ws.cell(row, 8, _r(det[0])); ws.cell(row, 9, _r(det[1])); ws.cell(row, 10, _r(det[2]))
        ws.cell(row, 11, _r(_duration(b, b['missile']), 4))
        ws.cell(row, 12, b['missile'])             # 干扰的导弹编号
    wb.save(out)
    return out
